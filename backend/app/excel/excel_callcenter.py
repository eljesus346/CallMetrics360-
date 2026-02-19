from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, date, timedelta
from collections import defaultdict


# ============================================================
# PALETA DE COLORES POR COLA - fondos suaves
# ============================================================
PALETA_9008 = {
    "dark":    "0F4C75",   # Azul petróleo
    "mid":     "1B6CA8",
    "light":   "EDF5FC",   # muy claro
    "lighter": "F7FBFF",   # casi blanco sutil
    "accent":  "4BA3E0",   # un poco más suave que el original
    "tab":     "1B6CA8",
}

# Paleta para hoja resumen (sin cambios)
PAL_RC = {
    "dark": "1A3C5E",
    "mid": "2E75B6",
    "accent": "41A0E0",
    "light": "D6E4F0",
    "lighter": "EBF3FA"
}

# Nombres completos para títulos internos (hojas por cola)
COLA_NOMBRES = {
    "9001": "IPS — Cola 9001",
    "9002": "Particulares — Cola 9002",
    "9003": "IPS — Cola 9003",
    "9004": "Robot — Cola 9004",
    "9007": "HUSJ — Cola 9007",
    "9008": "Rehabilitación — Cola 9008",
    "9011": "IPS — Cola 9011",
    "9014": "PAC — Cola 9014",
}

# *** NOMBRES ACTUALIZADOS PARA HOJA RESUMEN COLAS ***
COLA_NOMBRES_RESUMEN = {
    "9001": "9001 : Rehabilitar Imágenes",
    "9002": "9002 : Particulares",
    "9003": "9003 : IPS",
    "9004": "9004 : Robot",
    "9007": "9007 : Hospital San José",
    "9008": "9008 : Call Rehabilitar",
    "9011": "9011 : Preferencia Médicos",
    "9014": "9014 : PAC",
}

# Nombres cortos para pestañas
COLA_TABS = {
    "9002": "IPS-9002",
    "9003": "PARTICULARES-9003",
    "9004": "ROBOT-9004",
    "9007": "HUSJ-9007",
    "9008": "REHAB-9008",
    "9014": "PAC-9014",
    "9001": "IPS-9001",
    "9011": "IPS-9011",
}

DIAS_ES = {
    "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
    "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"
}
DIAS_UPPER = {
    "Monday": "LUNES", "Tuesday": "MARTES", "Wednesday": "MIÉRCOLES",
    "Thursday": "JUEVES", "Friday": "VIERNES", "Saturday": "SÁBADO", "Sunday": "DOMINGO"
}

MESES_ES = {
    "Jan": "Ene", "Feb": "Feb", "Mar": "Mar", "Apr": "Abr",
    "May": "May", "Jun": "Jun", "Jul": "Jul", "Aug": "Ago",
    "Sep": "Sep", "Oct": "Oct", "Nov": "Nov", "Dec": "Dic"
}


# ============================================================
# HELPERS
# ============================================================

def segundos_a_hhmmss(segundos):
    if not segundos:
        return "0:00:00 min"
    h = int(segundos) // 3600
    m = (int(segundos) % 3600) // 60
    s = int(segundos) % 60
    return f"{h}:{str(m).zfill(2)}:{str(s).zfill(2)} min"


def _mk_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _mk_border(color="D3D3D3", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _mk_border_thick_bottom(color_sides="D3D3D3", color_bottom="888888"):
    thin  = Side(style="thin",   color=color_sides)
    thick = Side(style="medium", color=color_bottom)
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def _c(ws, row, col, value=None, bg="FFFFFF", fg="000000", bold=False,
       size=10, halign="center", valign="center", wrap=False,
       border=None, italic=False, indent=0):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, italic=italic, color=fg, size=size, name="Arial")
    cell.fill = _mk_fill(bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               wrap_text=wrap, indent=indent)
    if border is not None:
        cell.border = border
    return cell


def _formato_fecha(fecha_str):
    dt = datetime.strptime(str(fecha_str), "%Y-%m-%d")
    mes = MESES_ES.get(dt.strftime("%b"), dt.strftime("%b"))
    return f"{dt.day:02d}-{mes}-{dt.year}"


def _formato_fecha_corto(fecha_str):
    dt = datetime.strptime(str(fecha_str), "%Y-%m-%d")
    mes = MESES_ES.get(dt.strftime("%b"), dt.strftime("%b"))
    return f"{dt.day}-{mes}"


# ============================================================
# AGRUPADOR DE SEMANAS
# ============================================================

def agrupar_por_semanas(data: list) -> list:
    if not data:
        return []
    data = [d for d in data
            if datetime.strptime(str(d["fecha"]), "%Y-%m-%d").weekday() != 6]
    if not data:
        return []

    data_sorted = sorted(data, key=lambda d: str(d["fecha"]))
    primera_fecha = datetime.strptime(str(data_sorted[0]["fecha"]), "%Y-%m-%d").date()
    dias_hasta_sabado = (5 - primera_fecha.weekday()) % 7
    fin_semana = primera_fecha + timedelta(days=dias_hasta_sabado)

    semanas, semana_actual = [], []
    for d in data_sorted:
        fecha = datetime.strptime(str(d["fecha"]), "%Y-%m-%d").date()
        if fecha <= fin_semana:
            semana_actual.append(d)
        else:
            if semana_actual:
                semanas.append(semana_actual)
            semana_actual = [d]
            dias_hasta_sabado = (5 - fecha.weekday()) % 7
            fin_semana = fecha + timedelta(days=dias_hasta_sabado)
    if semana_actual:
        semanas.append(semana_actual)
    return semanas


# ============================================================
# RENDER DE ENCABEZADO DE EMPRESA
# ============================================================

def _render_header_empresa(ws, cola: str, total_cols: int, row: int, pal: dict) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    border_h = _mk_border(pal["dark"], "thin")
    _c(ws, row, 1,
       value="🏥  IPS NUEVA POPAYÁN  —  Sistema de Reportes Call Center",
       bg=pal["dark"], fg="FFFFFF", bold=True, size=13,
       halign="center", border=border_h)
    ws.row_dimensions[row].height = 28
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill = _mk_fill(pal["dark"])
        ws.cell(row=row, column=col).border = border_h
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value=f"  {COLA_NOMBRES.get(cola, f'Cola {cola}')}",
       bg=pal["mid"], fg="FFFFFF", bold=True, size=11,
       halign="left", border=_mk_border(pal["mid"]))
    ws.row_dimensions[row].height = 22
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill = _mk_fill(pal["mid"])
        ws.cell(row=row, column=col).border = _mk_border(pal["mid"])
    row += 1

    ws.row_dimensions[row].height = 4
    for col in range(1, total_cols + 1):
        ws.cell(row=row, column=col).fill = _mk_fill(pal["accent"])
    row += 1

    return row


# ============================================================
# SEMÁFOROS EN TONOS PASTEL / SUAVES
# ============================================================

def semaforo_abandono(pct):
    if pct > 20:   return ("F9D8D8", "333333")
    if pct > 15:   return ("FFE8CC", "333333")
    if pct > 10:   return ("FFF4CC", "333333")
    return         ("D4EDDA", "333333")


def semaforo_respondidas(pct):
    if pct >= 90:  return ("D4EDDA", "333333")
    if pct >= 80:  return ("FFF4CC", "333333")
    if pct >= 70:  return ("FFE8CC", "333333")
    return         ("F9D8D8", "333333")


# ============================================================
# RENDER DE SEMANA POR COLA
# ============================================================

def _render_semana_cola(ws, data_semanal: list, start_row: int,
                        semana_num: int, cola: str, pal: dict) -> int:

    nombre_cola = COLA_NOMBRES.get(cola, f"Cola {cola}")
    total_cols  = 1 + len(data_semanal) * 2 + 2
    row         = start_row

    border_dark  = _mk_border(pal["dark"])
    border_mid   = _mk_border(pal["mid"])
    border_light = _mk_border("CCCCCC")

    fecha_min = _formato_fecha(min(d["fecha"] for d in data_semanal))
    fecha_max = _formato_fecha(max(d["fecha"] for d in data_semanal))

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value=f"  SEMANA {semana_num}   ·   {nombre_cola}   ·   {fecha_min}  →  {fecha_max}",
       bg=pal["dark"], fg="FFFFFF", bold=True, size=11,
       halign="left", border=border_dark)
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill = _mk_fill(pal["dark"])
        ws.cell(row=row, column=col).border = border_dark
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    _c(ws, row, 1, value="MÉTRICA", bg=pal["mid"], fg="FFFFFF",
       bold=True, size=9, halign="center", border=border_mid)
    ws.cell(row=row + 1, column=1).fill   = _mk_fill(pal["mid"])
    ws.cell(row=row + 1, column=1).border = border_mid

    col = 2
    for d in data_semanal:
        dt = datetime.strptime(str(d["fecha"]), "%Y-%m-%d")
        nombre_dia = DIAS_UPPER.get(dt.strftime("%A"), dt.strftime("%A"))
        mes = MESES_ES.get(dt.strftime("%b"), dt.strftime("%b"))

        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        _c(ws, row, col, value=nombre_dia,
           bg=pal["mid"], fg="FFFFFF", bold=True, size=10,
           halign="center", border=border_mid)
        ws.cell(row=row, column=col + 1).fill   = _mk_fill(pal["mid"])
        ws.cell(row=row, column=col + 1).border = border_mid

        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        _c(ws, row + 1, col, value=f"{dt.day}-{mes}",
           bg=pal["accent"], fg="FFFFFF", bold=False, size=9,
           halign="center", border=border_mid)
        ws.cell(row=row + 1, column=col + 1).fill   = _mk_fill(pal["accent"])
        ws.cell(row=row + 1, column=col + 1).border = border_mid
        col += 2

    ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + 1)
    _c(ws, row, col, value="TOTAL\nSEMANA",
       bg=pal["dark"], fg="FFFFFF", bold=True, size=10,
       halign="center", wrap=True, border=border_dark)
    for c in range(col, col + 2):
        for r in range(row, row + 2):
            ws.cell(row=r, column=c).fill   = _mk_fill(pal["dark"])
            ws.cell(row=r, column=c).border = border_dark

    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 16
    row += 2

    total_llamadas    = sum((d.get("llamadas_totales", 0) or 0) for d in data_semanal)
    total_respondidas = sum((d.get("respondidas", 0) or 0)      for d in data_semanal)
    total_abandonadas = sum((d.get("abandonadas", 0) or 0)      for d in data_semanal)
    pct_abandon_total = round(total_abandonadas / total_llamadas * 100, 1) if total_llamadas > 0 else 0

    SECCION_VOLUMETRIA = "VOLUMEN DE LLAMADAS"
    SECCION_TRAFICO    = "DISTRIBUCIÓN DE TRÁFICO"
    SECCION_TIEMPOS    = "TIEMPOS"

    filas = [
        (SECCION_VOLUMETRIA, None, None, None, False, True),
        ("Llamadas Total", "llamadas", lambda d: d.get("llamadas_totales", 0) or 0, lambda d: "", True, False),
        ("Respondidas", "respondidas", lambda d: d.get("respondidas", 0) or 0, lambda d: f'{d.get("pct_respondidas", 0) or 0}%', False, False),
        ("Abandonadas", "abandonadas", lambda d: d.get("abandonadas", 0) or 0, lambda d: f'{d.get("pct_abandonadas", 0) or 0}%', False, False),
        (SECCION_TRAFICO, None, None, None, False, True),
        ("Horario + Tráfico", "horario", lambda d: d.get("hora_pico") or "-", lambda d: f'{d.get("cantidad_hora_pico", 0) or 0} Llam' if d.get("cantidad_hora_pico") else "-", False, False),
        ("Horario - Tráfico", "horario", lambda d: d.get("hora_menos_pico") or "-", lambda d: f'{d.get("cantidad_hora_menos_pico", 0) or 0} Llam' if d.get("cantidad_hora_menos_pico") else "-", False, False),
        (SECCION_TIEMPOS, None, None, None, False, True),
        ("Promedio Espera", "tiempo", lambda d: f'{d.get("promedio_espera") or 0} s', lambda d: f'{round((d.get("promedio_espera") or 0) / 60)} min', False, False),
        ("Espera + Larga", "tiempo", lambda d: f'{d.get("espera_mas_larga") or 0} s', lambda d: f'{round((d.get("espera_mas_larga") or 0) / 60)} min', False, False),
        ("Prom. Dur. Llamada", "duracion", lambda d: "", lambda d: segundos_a_hhmmss(d.get("promedio_duracion_llamada") or 0), True, False),
        ("Dur. + Larga Llamada", "duracion", lambda d: "", lambda d: segundos_a_hhmmss(d.get("duracion_mas_larga") or 0), True, False),
    ]

    for fila_idx, (titulo, tipo, fn1, fn2, combinar, es_seccion) in enumerate(filas):

        if es_seccion:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
            _c(ws, row, 1, value=f"  ▸  {titulo}",
               bg=pal["dark"], fg="FFFFFF", bold=True, size=9,
               halign="left", border=border_dark)
            for col in range(2, total_cols + 1):
                ws.cell(row=row, column=col).fill   = _mk_fill(pal["dark"])
                ws.cell(row=row, column=col).border = border_dark
            ws.row_dimensions[row].height = 14
            row += 1
            continue

        dato_idx = sum(1 for f in filas[:fila_idx] if not f[5])
        bg_fila  = pal["light"] if dato_idx % 2 == 0 else pal["lighter"]

        _c(ws, row, 1, value=titulo,
           bg=bg_fila, fg=pal["dark"], bold=True, size=10,
           halign="left", indent=1,
           border=_mk_border_thick_bottom("CCCCCC", pal["mid"]))
        ws.row_dimensions[row].height = 18

        col = 2
        for d in data_semanal:
            v1 = fn1(d)
            v2 = fn2(d)

            if combinar:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                _c(ws, row, col, value=v2 if not v1 else v1,
                   bg=bg_fila, fg="333333", size=10,
                   halign="center", border=border_light)
                ws.cell(row=row, column=col + 1).fill   = _mk_fill(bg_fila)
                ws.cell(row=row, column=col + 1).border = border_light
            else:
                bg_v2 = bg_fila
                fg_v2 = "333333"
                if tipo == "respondidas" and v2:
                    pct_val = float(str(v2).replace("%", "") or 0)
                    bg_v2, fg_v2 = semaforo_respondidas(pct_val)
                elif tipo == "abandonadas" and v2:
                    pct_val = float(str(v2).replace("%", "") or 0)
                    bg_v2, fg_v2 = semaforo_abandono(pct_val)

                _c(ws, row, col,     value=v1, bg=bg_fila, fg="333333",
                   size=10, halign="center", border=border_light)
                _c(ws, row, col + 1, value=v2, bg=bg_v2,   fg=fg_v2,
                   bold=(tipo in ("respondidas", "abandonadas")),
                   size=10, halign="center", border=border_light)
            col += 2

        col_total = col
        if tipo == "llamadas":
            ws.merge_cells(start_row=row, start_column=col_total, end_row=row, end_column=col_total + 1)
            _c(ws, row, col_total, value=total_llamadas,
               bg=pal["mid"], fg="FFFFFF", bold=True, size=11,
               halign="center", border=border_mid)
            ws.cell(row=row, column=col_total + 1).fill   = _mk_fill(pal["mid"])
            ws.cell(row=row, column=col_total + 1).border = border_mid

        elif tipo == "respondidas":
            pct = round(total_respondidas / total_llamadas * 100, 1) if total_llamadas > 0 else 0
            bg_tot, fg_tot = semaforo_respondidas(pct)
            _c(ws, row, col_total,     value=total_respondidas,
               bg=pal["lighter"], fg=pal["dark"], bold=True, size=10,
               halign="center", border=border_mid)
            _c(ws, row, col_total + 1, value=f"{pct}%",
               bg=bg_tot, fg=fg_tot, bold=True, size=10,
               halign="center", border=border_mid)

        elif tipo == "abandonadas":
            bg_tot, fg_tot = semaforo_abandono(pct_abandon_total)
            _c(ws, row, col_total,     value=total_abandonadas,
               bg=pal["lighter"], fg=pal["dark"], bold=True, size=10,
               halign="center", border=border_mid)
            _c(ws, row, col_total + 1, value=f"{pct_abandon_total}%",
               bg=bg_tot, fg=fg_tot, bold=True, size=10,
               halign="center", border=border_mid)

        else:
            ws.merge_cells(start_row=row, start_column=col_total, end_row=row, end_column=col_total + 1)
            _c(ws, row, col_total, value="",
               bg=pal["lighter"], fg="000000", border=border_light)
            ws.cell(row=row, column=col_total + 1).fill   = _mk_fill(pal["lighter"])
            ws.cell(row=row, column=col_total + 1).border = border_light

        row += 1

    for col in range(1, total_cols + 1):
        ws.cell(row=row, column=col).fill   = _mk_fill(pal["accent"])
        ws.cell(row=row, column=col).border = _mk_border(pal["accent"])
    ws.row_dimensions[row].height = 3
    row += 1

    return row + 2


# ============================================================
# SEMÁFOROS PASTEL PARA HOJA RESUMEN
# ============================================================

def sem_aband_rc(pct):
    if pct > 20:  return ("F9D8D8", "333333")
    if pct > 15:  return ("FFE8CC", "333333")
    if pct > 10:  return ("FFF4CC", "333333")
    return        ("D4EDDA", "333333")


def sem_exito_rc(pct):
    if pct >= 90: return ("D4EDDA", "333333")
    if pct >= 80: return ("FFF4CC", "333333")
    if pct >= 70: return ("FFE8CC", "333333")
    return        ("F9D8D8", "333333")


# ============================================================
# RENDER HOJA RESUMEN COLAS
# ============================================================

def _render_hoja2_semana(ws2, data_por_cola: list, start_row: int,
                         semana_num: int, fecha_min: str, fecha_max: str,
                         colas_globales: list = None) -> int:

    datos_por_cola    = defaultdict(dict)
    for d in data_por_cola:
        datos_por_cola[d["cola"]][str(d["fecha"])] = d

    # *** ORDENAR COLAS DE MENOR A MAYOR (numéricamente) ***
    if colas_globales:
        colas_ordenadas = sorted(colas_globales, key=lambda c: int(c))
    else:
        colas_ordenadas = sorted(
            datos_por_cola.keys(),
            key=lambda c: int(c)
        )

    fechas_ordenadas = sorted(set(str(d["fecha"]) for d in data_por_cola))

    COLS_POR_DIA  = 5
    total_cols_h2 = 1 + len(fechas_ordenadas) * COLS_POR_DIA

    row = start_row
    bord = _mk_border("CCCCCC")

    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols_h2)
    _c(ws2, row, 1,
       value="🏥  IPS NUEVA POPAYÁN  —  Resumen de Comportamiento por Cola",
       bg=PAL_RC["dark"], fg="FFFFFF", bold=True, size=12,
       halign="center", border=_mk_border(PAL_RC["dark"]))
    for col in range(2, total_cols_h2 + 1):
        ws2.cell(row=row, column=col).fill   = _mk_fill(PAL_RC["dark"])
        ws2.cell(row=row, column=col).border = _mk_border(PAL_RC["dark"])
    ws2.row_dimensions[row].height = 26
    row += 1

    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols_h2)
    _c(ws2, row, 1,
       value=f"  SEMANA {semana_num}   ·   {_formato_fecha(fecha_min)}  →  {_formato_fecha(fecha_max)}",
       bg=PAL_RC["mid"], fg="FFFFFF", bold=True, size=11,
       halign="left", border=_mk_border(PAL_RC["mid"]))
    for col in range(2, total_cols_h2 + 1):
        ws2.cell(row=row, column=col).fill   = _mk_fill(PAL_RC["mid"])
        ws2.cell(row=row, column=col).border = _mk_border(PAL_RC["mid"])
    ws2.row_dimensions[row].height = 20
    row += 1

    for col in range(1, total_cols_h2 + 1):
        ws2.cell(row=row, column=col).fill = _mk_fill(PAL_RC["accent"])
    ws2.row_dimensions[row].height = 3
    row += 1

    _c(ws2, row, 1, value="", bg="FFFFFF", border=bord)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    ws2.cell(row=row + 1, column=1).fill   = _mk_fill("FFFFFF")
    ws2.cell(row=row + 1, column=1).border = bord

    col = 2
    for fecha in fechas_ordenadas:
        dt     = datetime.strptime(fecha, "%Y-%m-%d")
        dia_es = DIAS_ES.get(dt.strftime("%A"), dt.strftime("%A"))
        mes    = MESES_ES.get(dt.strftime("%b"), dt.strftime("%b"))

        ws2.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + COLS_POR_DIA - 1)
        _c(ws2, row, col, value=f"{dt.day}-{mes}", bg=PAL_RC["mid"],
           fg="FFFFFF", bold=True, size=10, halign="center", border=_mk_border(PAL_RC["mid"]))
        for c2 in range(col + 1, col + COLS_POR_DIA):
            ws2.cell(row=row, column=c2).fill   = _mk_fill(PAL_RC["mid"])
            ws2.cell(row=row, column=c2).border = _mk_border(PAL_RC["mid"])

        ws2.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + COLS_POR_DIA - 1)
        _c(ws2, row + 1, col, value=dia_es, bg=PAL_RC["accent"],
           fg="FFFFFF", bold=False, size=10, halign="center", border=_mk_border(PAL_RC["accent"]))
        for c2 in range(col + 1, col + COLS_POR_DIA):
            ws2.cell(row=row + 1, column=c2).fill   = _mk_fill(PAL_RC["accent"])
            ws2.cell(row=row + 1, column=c2).border = _mk_border(PAL_RC["accent"])
        col += COLS_POR_DIA

    ws2.row_dimensions[row].height = 18
    ws2.row_dimensions[row + 1].height = 16
    row += 2

    _c(ws2, row, 1, value="COLA", bg=PAL_RC["dark"], fg="FFFFFF",
       bold=True, size=10, halign="center", border=_mk_border(PAL_RC["dark"]))
    col = 2
    for _ in fechas_ordenadas:
        for lbl, w_bg in [("Exitosas", PAL_RC["dark"]), ("%", PAL_RC["dark"]),
                          ("Aband.", PAL_RC["dark"]),   ("%", PAL_RC["dark"]),
                          ("Total",  PAL_RC["dark"])]:
            _c(ws2, row, col, value=lbl, bg=w_bg, fg="FFFFFF",
               bold=True, size=9, halign="center", border=_mk_border(PAL_RC["dark"]))
            col += 1
    ws2.row_dimensions[row].height = 18
    row += 1

    for i, cola in enumerate(colas_ordenadas):
        bg = PAL_RC["light"] if i % 2 == 0 else PAL_RC["lighter"]

        # *** USAR NOMBRE ACTUALIZADO PARA HOJA RESUMEN ***
        nombre_cola = COLA_NOMBRES_RESUMEN.get(cola, f"{cola} : Cola {cola}")
        _c(ws2, row, 1, value=nombre_cola, bg=PAL_RC["mid"], fg="FFFFFF",
           bold=True, size=9, halign="left", indent=1, border=_mk_border(PAL_RC["mid"]))

        col = 2
        for fecha in fechas_ordenadas:
            d           = datos_por_cola[cola].get(fecha, {})
            total       = d.get("llamadas_totales", 0) or 0
            respondidas = d.get("respondidas", 0) or 0
            abandonadas = d.get("abandonadas", 0) or 0
            pct_exito   = round(respondidas / total * 100) if total > 0 else 0
            pct_aband   = round(abandonadas  / total * 100) if total > 0 else 0

            bg_ex, fg_ex = sem_exito_rc(pct_exito)
            bg_ab, fg_ab = sem_aband_rc(pct_aband)

            _c(ws2, row, col,     value=respondidas,          bg=bg,    fg="333333", size=9, halign="center", border=bord)
            _c(ws2, row, col + 1, value=f"{int(pct_exito)}%", bg=bg_ex, fg=fg_ex,   size=9, bold=True, halign="center", border=bord)
            _c(ws2, row, col + 2, value=abandonadas,          bg=bg,    fg="333333", size=9, halign="center", border=bord)
            _c(ws2, row, col + 3, value=f"{int(pct_aband)}%", bg=bg_ab, fg=fg_ab,   size=9, bold=True, halign="center", border=bord)
            _c(ws2, row, col + 4, value=total,                bg=bg,    fg=PAL_RC["dark"], bold=True, size=9, halign="center", border=bord)
            col += COLS_POR_DIA

        ws2.row_dimensions[row].height = 17
        row += 1

    _c(ws2, row, 1, value="▸  TOTAL", bg=PAL_RC["dark"], fg="FFFFFF",
       bold=True, size=10, halign="left", indent=1, border=_mk_border(PAL_RC["dark"]))
    col = 2
    for fecha in fechas_ordenadas:
        tot     = sum((d.get("llamadas_totales", 0) or 0) for d in data_por_cola if str(d["fecha"]) == fecha)
        res     = sum((d.get("respondidas", 0)      or 0) for d in data_por_cola if str(d["fecha"]) == fecha)
        abd     = sum((d.get("abandonadas", 0)       or 0) for d in data_por_cola if str(d["fecha"]) == fecha)
        pct_res = round(res / tot * 100) if tot > 0 else 0
        pct_abd = round(abd / tot * 100) if tot > 0 else 0

        bg_ex, fg_ex = sem_exito_rc(pct_res)
        bg_ab, fg_ab = sem_aband_rc(pct_abd)

        _c(ws2, row, col,     value=res,            bg=PAL_RC["mid"],  fg="FFFFFF", bold=True, size=10, halign="center", border=_mk_border(PAL_RC["mid"]))
        _c(ws2, row, col + 1, value=f"{pct_res}%",  bg=bg_ex,           fg=fg_ex,   bold=True, size=10, halign="center", border=_mk_border(PAL_RC["mid"]))
        _c(ws2, row, col + 2, value=abd,             bg=PAL_RC["mid"],  fg="FFFFFF", bold=True, size=10, halign="center", border=_mk_border(PAL_RC["mid"]))
        _c(ws2, row, col + 3, value=f"{pct_abd}%",  bg=bg_ab,           fg=fg_ab,   bold=True, size=10, halign="center", border=_mk_border(PAL_RC["mid"]))
        _c(ws2, row, col + 4, value=tot,             bg=PAL_RC["dark"], fg="FFFFFF", bold=True, size=10, halign="center", border=_mk_border(PAL_RC["dark"]))
        col += COLS_POR_DIA

    ws2.row_dimensions[row].height = 20
    row += 1

    # *** COLUMNA A MÁS ANCHA PARA VER NOMBRES COMPLETOS ***
    ws2.column_dimensions["A"].width = 34
    col = 2
    for _ in fechas_ordenadas:
        for offset, w in enumerate([9, 7, 9, 7, 9]):
            ws2.column_dimensions[get_column_letter(col + offset)].width = w
        col += COLS_POR_DIA

    return row + 3


# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def generar_excel_callcenter(data_por_cola: list, data_detalle_por_cola: dict):
    output = BytesIO()
    wb = Workbook()
    first_sheet_used = False

    totales_por_cola = defaultdict(int)
    for d in data_por_cola:
        totales_por_cola[d["cola"]] += d.get("llamadas_totales", 0) or 0

    # Orden de pestañas por cola (mayor a menor volumen, como estaba)
    colas_globales = sorted(totales_por_cola.keys(),
                             key=lambda c: totales_por_cola[c], reverse=True)

    # HOJAS POR COLA
    for cola in colas_globales:
        data_cola   = data_detalle_por_cola.get(cola, [])
        tab_name    = COLA_TABS.get(cola, cola)
        pal         = PALETA_9008.copy()

        if not first_sheet_used:
            ws = wb.active
            ws.title = tab_name
            first_sheet_used = True
        else:
            ws = wb.create_sheet(title=tab_name)

        ws.sheet_properties.tabColor = pal["tab"]

        if not data_cola:
            ws["A1"] = f"No hay datos para la cola {cola}"
            continue

        semanas = agrupar_por_semanas(data_cola)

        max_dias    = max(len(s) for s in semanas) if semanas else 0
        total_cols  = 1 + max_dias * 2 + 2

        current_row = _render_header_empresa(ws, cola, total_cols, 1, pal)

        for i, semana_data in enumerate(semanas, start=1):
            current_row = _render_semana_cola(ws, semana_data, current_row, i, cola, pal)

        _c(ws, current_row, 1,
           value=f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Reportes Call Center",
           bg="FFFFFF", fg="999999", italic=True, size=8,
           halign="left", border=_mk_border("FFFFFF"))
        ws.row_dimensions[current_row].height = 16

        max_total_cols = max(1 + len(s) * 2 + 2 for s in semanas) if semanas else 10
        ws.column_dimensions["A"].width = 24
        for col_idx in range(2, max_total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13

        ws.freeze_panes = "B1"
        ws.sheet_view.showGridLines = False

    # HOJA RESUMEN COLAS
    ws2 = wb.create_sheet(title="Resumen Colas")
    ws2.sheet_properties.tabColor = "1A3C5E"
    ws2.sheet_view.showGridLines = False

    if not data_por_cola:
        ws2["A1"] = "No hay datos por cola"
    else:
        semanas_h2 = agrupar_por_semanas(data_por_cola)
        current_row = 1
        for i, semana_data in enumerate(semanas_h2, start=1):
            fecha_min_s = min(str(d["fecha"]) for d in semana_data)
            fecha_max_s = max(str(d["fecha"]) for d in semana_data)
            current_row = _render_hoja2_semana(
                ws2, semana_data, current_row, i,
                fecha_min_s, fecha_max_s,
                colas_globales=colas_globales
            )
        ws2.freeze_panes = "B1"

    wb.save(output)
    output.seek(0)
    return output