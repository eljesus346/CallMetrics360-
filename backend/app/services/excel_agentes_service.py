from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


# ============================================================
# PALETA CORPORATIVA — misma base que Call Center
# ============================================================
PAL = {
    "dark":    "1A3C5E",   # Azul oscuro — encabezados principales
    "mid":     "2E75B6",   # Azul medio — subencabezados
    "accent":  "41A0E0",   # Azul claro — separadores / accent
    "light":   "D6E4F0",   # Fila alterna clara
    "lighter": "EBF3FA",   # Fila alterna más clara
    "tab":     "1A3C5E",
}

# Semáforos pastel — mismos criterios que Call Center
def _sem_estado(estado: str):
    e = (estado or "").strip().lower()
    if e == "disponible":  return ("D4EDDA", "333333")   # verde pastel
    if e == "pausa":       return ("FFF4CC", "333333")   # amarillo pastel
    if e == "desconectado": return ("F9D8D8", "333333")  # rojo pastel
    return ("FFFFFF", "333333")


# ============================================================
# HELPERS — reutilizados del módulo Call Center
# ============================================================

def _mk_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _mk_border(color="D3D3D3", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _mk_border_thick_bottom(color_sides="D3D3D3", color_bottom="888888"):
    thin  = Side(style="thin",   color=color_sides)
    thick = Side(style="medium", color=color_bottom)
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def _c(ws, row, col, value=None, bg="FFFFFF", fg="000000", bold=False,
       size=10, halign="center", valign="center", wrap=False,
       border=None, italic=False, indent=0):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font      = Font(bold=bold, italic=italic, color=fg, size=size, name="Arial")
    cell.fill      = _mk_fill(bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                wrap_text=wrap, indent=indent)
    if border is not None:
        cell.border = border
    return cell


def segundos_a_hhmmss(segundos):
    if not segundos:
        return "0:00:00"
    h = int(segundos) // 3600
    m = (int(segundos) % 3600) // 60
    s = int(segundos) % 60
    return f"{h}:{str(m).zfill(2)}:{str(s).zfill(2)}"


# ============================================================
# ENCABEZADO CORPORATIVO — igual que Call Center
# ============================================================

def _render_header(ws, total_cols: int, fecha_str: str) -> int:
    row = 1
    border_dark = _mk_border(PAL["dark"], "thin")

    # Fila 1 — Nombre empresa
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value="🏥  IPS NUEVA POPAYÁN  —  Sistema de Reportes Call Center",
       bg=PAL["dark"], fg="FFFFFF", bold=True, size=13,
       halign="center", border=border_dark)
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill   = _mk_fill(PAL["dark"])
        ws.cell(row=row, column=col).border = border_dark
    ws.row_dimensions[row].height = 28
    row += 1

    # Fila 2 — Subtítulo del reporte
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value="  Control de Agentes  —  Tiempos, Pausas y Cumplimiento",
       bg=PAL["mid"], fg="FFFFFF", bold=True, size=11,
       halign="left", border=_mk_border(PAL["mid"]))
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill   = _mk_fill(PAL["mid"])
        ws.cell(row=row, column=col).border = _mk_border(PAL["mid"])
    ws.row_dimensions[row].height = 22
    row += 1

    # Fila 3 — Fecha generación
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value=f"  Generado el: {fecha_str}",
       bg=PAL["lighter"], fg=PAL["dark"], italic=True, size=9,
       halign="left", border=_mk_border("CCCCCC"))
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill   = _mk_fill(PAL["lighter"])
        ws.cell(row=row, column=col).border = _mk_border("CCCCCC")
    ws.row_dimensions[row].height = 16
    row += 1

    # Fila 4 — Separador accent
    for col in range(1, total_cols + 1):
        ws.cell(row=row, column=col).fill = _mk_fill(PAL["accent"])
    ws.row_dimensions[row].height = 4
    row += 1

    return row


# ============================================================
# ENCABEZADOS DE COLUMNAS
# ============================================================

COLUMNAS = [
    ("Agente",           45),
    ("Tiempo Logueado",  18),
    ("Tiempo Activo",    18),
    ("Tipo Pausa",       22),
    ("Tiempo Pausa",     18),
    ("Estado",           16),
]

def _render_col_headers(ws, row: int) -> int:
    border_mid = _mk_border(PAL["mid"])

    # Fila encabezados
    for col_idx, (nombre, _) in enumerate(COLUMNAS, start=1):
        _c(ws, row, col_idx,
           value=nombre,
           bg=PAL["mid"], fg="FFFFFF", bold=True, size=10,
           halign="center", border=border_mid)
    ws.row_dimensions[row].height = 22
    row += 1

    # Separador accent delgado
    for col_idx in range(1, len(COLUMNAS) + 1):
        ws.cell(row=row, column=col_idx).fill   = _mk_fill(PAL["accent"])
        ws.cell(row=row, column=col_idx).border = _mk_border(PAL["accent"])
    ws.row_dimensions[row].height = 3
    row += 1

    return row


# ============================================================
# FILAS DE DATOS
# ============================================================

def _render_datos(ws, data: list, start_row: int) -> int:
    row         = start_row
    border_data = _mk_border("CCCCCC")

    for i, d in enumerate(data):
        bg_base = PAL["light"] if i % 2 == 0 else PAL["lighter"]

        estado           = d.get("estado", "") or ""
        bg_estado, fg_est = _sem_estado(estado)

        valores = [
            d.get("agente", ""),
            segundos_a_hhmmss(d.get("tiempo_logueado")),
            segundos_a_hhmmss(d.get("tiempo_activo")),
            d.get("tipo_pausa", "") or "",
            segundos_a_hhmmss(d.get("tiempo_pausa")),
            estado,
        ]

        for col_idx, val in enumerate(valores, start=1):
            if col_idx == 6:   # Estado — semáforo
                _c(ws, row, col_idx, value=val,
                   bg=bg_estado, fg=fg_est, bold=True, size=10,
                   halign="center",
                   border=_mk_border_thick_bottom("CCCCCC", PAL["mid"]))
            elif col_idx == 1:  # Agente — alineado a la izquierda
                _c(ws, row, col_idx, value=val,
                   bg=bg_base, fg=PAL["dark"], bold=True, size=10,
                   halign="left", indent=1,
                   border=_mk_border_thick_bottom("CCCCCC", PAL["mid"]))
            else:
                _c(ws, row, col_idx, value=val,
                   bg=bg_base, fg="333333", size=10,
                   halign="center", border=border_data)

        ws.row_dimensions[row].height = 18
        row += 1

    return row


# ============================================================
# FILA TOTALES / RESUMEN
# ============================================================

def _render_totales(ws, data: list, row: int) -> int:
    total_cols = len(COLUMNAS)
    border_dark = _mk_border(PAL["dark"])

    disponibles   = sum(1 for d in data if (d.get("estado") or "").strip().lower() == "disponible")
    en_pausa      = sum(1 for d in data if (d.get("estado") or "").strip().lower() == "pausa")
    desconectados = sum(1 for d in data if (d.get("estado") or "").strip().lower() == "desconectado")
    total_agentes = len(data)

    # Separador
    for col in range(1, total_cols + 1):
        ws.cell(row=row, column=col).fill   = _mk_fill(PAL["accent"])
        ws.cell(row=row, column=col).border = _mk_border(PAL["accent"])
    ws.row_dimensions[row].height = 3
    row += 1

    # Encabezado fila resumen
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value="  ▸  RESUMEN DE ESTADO DE AGENTES",
       bg=PAL["dark"], fg="FFFFFF", bold=True, size=9,
       halign="left", border=border_dark)
    for col in range(2, total_cols + 1):
        ws.cell(row=row, column=col).fill   = _mk_fill(PAL["dark"])
        ws.cell(row=row, column=col).border = border_dark
    ws.row_dimensions[row].height = 14
    row += 1

    resumen = [
        ("Total Agentes",  total_agentes, PAL["mid"],    "FFFFFF"),
        ("Disponibles",    disponibles,   "D4EDDA",      "333333"),
        ("En Pausa",       en_pausa,      "FFF4CC",      "333333"),
        ("Desconectados",  desconectados, "F9D8D8",      "333333"),
    ]

    col = 1
    for label, valor, bg, fg in resumen:
        _c(ws, row, col,     value=label, bg=PAL["lighter"], fg=PAL["dark"],
           bold=True, size=9, halign="left", indent=1,
           border=_mk_border("CCCCCC"))
        _c(ws, row, col + 1, value=valor, bg=bg, fg=fg,
           bold=True, size=11, halign="center",
           border=_mk_border("CCCCCC"))
        col += 2

    # Si hay columna sobrante (6 columnas, 4 pares = 8 → hay 2 extra)
    # llenamos el resto en blanco
    for c in range(col, total_cols + 1):
        _c(ws, row, c, bg=PAL["lighter"], border=_mk_border("CCCCCC"))

    ws.row_dimensions[row].height = 20
    row += 1

    return row


# ============================================================
# PIE DE PÁGINA
# ============================================================

def _render_footer(ws, row: int, total_cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _c(ws, row, 1,
       value=f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Reportes Call Center  |  IPS Nueva Popayán",
       bg="FFFFFF", fg="999999", italic=True, size=8,
       halign="left", border=_mk_border("FFFFFF"))
    ws.row_dimensions[row].height = 16


# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def generar_excel_agentes(data: list):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Control Agentes"
    ws.sheet_properties.tabColor  = PAL["tab"]
    ws.sheet_view.showGridLines   = False

    total_cols  = len(COLUMNAS)
    fecha_str   = datetime.now().strftime("%d/%m/%Y %H:%M")

    # 1. Encabezado corporativo
    current_row = _render_header(ws, total_cols, fecha_str)

    # 2. Encabezados de columnas
    current_row = _render_col_headers(ws, current_row)

    # 3. Datos
    if data:
        current_row = _render_datos(ws, data, current_row)
    else:
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        _c(ws, current_row, 1,
           value="No hay datos disponibles para el período seleccionado.",
           bg=PAL["lighter"], fg="888888", italic=True, size=10,
           halign="center", border=_mk_border("CCCCCC"))
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # 4. Pie de página
    _render_footer(ws, current_row, total_cols)

    # 6. Anchos de columnas
    for col_idx, (_, width) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 7. Congelar primera fila de datos
    ws.freeze_panes = "A7"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output